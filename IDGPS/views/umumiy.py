from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from IDGPS.models import Bugalteriya, Sklad, Sotish, Sklad_sotish
from django.views import View
from django.views.generic import TemplateView
from django.contrib import messages
from django.http import HttpResponse
from django.db import models
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import pandas as pd
import openpyxl


class MijozlarView(LoginRequiredMixin, TemplateView):
    template_name = "mijozlar.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Barcha Sotish yozuvlarini sanaga ko'ra tartiblash
        # Sklad_sotish ma'lumotlarini prefetch_related bilan olish
        mijozlar = (
            Sotish.objects.all()
            .order_by("-sana")
            .prefetch_related("sklad_sotish_set", "sklad_sotish_set__sklad")
        )

        mijozlar_data = []
        unique_mijozlar = set()

        for mijoz in mijozlar:
            # Mijoz nomini unique set ga qo'shish
            unique_mijozlar.add(mijoz.mijoz)

            # Aloqador GPS yozuvlarini olish, prefetch_related tufayli yangi query bajarilmaydi
            sklad_sotish_items = list(mijoz.sklad_sotish_set.all())

            # Har bir GPS va SIM karta uchun qator yaratish
            for i, sklad_sotish in enumerate(sklad_sotish_items):
                jami_summa = (
                    mijoz.summasi
                    + mijoz.master_summasi
                    + (mijoz.abonent_tulov * len(sklad_sotish_items))
                )

                mijozlar_data.append(
                    {
                        "id": mijoz.id,
                        "mijoz": mijoz.mijoz if i == 0 else "",
                        "tel_raqam": mijoz.mijoz_tel_raqam if i == 0 else "",
                        "username": mijoz.username if i == 0 else "",
                        "password": mijoz.password if i == 0 else "",
                        "dasturiy_taminot_nomi": (
                            str(mijoz.dasturiy_taminot)
                            if mijoz.dasturiy_taminot and i == 0
                            else ""
                        ),
                        "abonent_tulov": mijoz.abonent_tulov if i == 0 else "",
                        "sana": (
                            mijoz.sana.strftime("%Y-%m-%d")
                            if mijoz.sana and i == 0
                            else ""
                        ),
                        "summasi": mijoz.summasi if i == 0 else "",
                        "naqd": mijoz.naqd if i == 0 else "",
                        "bank_schot": mijoz.bank_schot if i == 0 else "",
                        "karta": mijoz.karta if i == 0 else "",
                        "master": mijoz.master if i == 0 else "",
                        "master_summasi": mijoz.master_summasi if i == 0 else "",
                        "gps": sklad_sotish.sklad.gps_id,
                        "sim_karta": sklad_sotish.sim_karta,
                        "mashina_turi": sklad_sotish.mashina_turi,
                        "davlat_raqami": sklad_sotish.davlat_raqami,
                        "rowspan": len(sklad_sotish_items) if i == 0 else 0,
                        "first_row": i == 0,
                        "jami_summa": jami_summa if i == 0 else "",
                    }
                )

        # Pagination implementation
        page = self.request.GET.get("page", 1)
        paginator = Paginator(mijozlar_data, 20)  # 20 items per page

        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        context["mijozlar_data"] = mijozlar_data  # For mobile view
        context["page_obj"] = page_obj  # For desktop view with pagination
        context["unique_mijoz_count"] = len(
            unique_mijozlar
        )  # Count of unique customers

        return context


class StatistikaView(LoginRequiredMixin, TemplateView):
    template_name = "statistika.html"
    oylar = [
        "Yanvar",
        "Fevral",
        "Mart",
        "Aprel",
        "May",
        "Iyun",
        "Iyul",
        "Avgust",
        "Sentabr",
        "Oktabr",
        "Noyabr",
        "Dekabr",
    ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Asosiy vaqt parametrlarini tayyorlash
        now = timezone.now()
        selected_year = int(self.request.GET.get("year", now.year))
        current_month = now.month if selected_year == now.year else 12

        # Barcha oylar uchun statistikani bir so'rov bilan olish
        monthly_stats = self.get_optimized_stats(selected_year, current_month)

        # Grafik va diagrammalar uchun qo'shimcha ma'lumotlar
        context.update(
            {
                "monthly_stats": monthly_stats,
                "current_year": selected_year,
                "current_month": now.month,
                "yillar": range(2020, now.year + 1),
                "chart_data": self.prepare_chart_data(monthly_stats, selected_year, current_month),
            }
        )

        return context

    def prepare_chart_data(self, monthly_stats, selected_year, current_month):
        """
        Grafiklar uchun qo'shimcha ma'lumotlarni tayyorlash
        """
        # Oylik o'sish foizlari
        growth_rates = {}
        revenue_growth = {}
        gps_penetration = {}
        payment_ratios = {}
        
        for month in range(1, current_month + 1):
            stats = monthly_stats.get(month, {})
            prev_month = month - 1
            
            # Abonentlar o'sishi
            if month > 1 and prev_month in monthly_stats:
                prev_count = monthly_stats[prev_month]["abonent"]["jami_aktiv"]
                current_count = stats["abonent"]["jami_aktiv"]
                if prev_count > 0:
                    growth_pct = ((current_count - prev_count) / prev_count) * 100
                else:
                    growth_pct = 100 if current_count > 0 else 0
                growth_rates[month] = round(growth_pct, 1)
            else:
                growth_rates[month] = 0
            
            # Tushum o'sishi
            if month > 1 and prev_month in monthly_stats:
                prev_revenue = monthly_stats[prev_month]["summa"]["oylik_umumiy_summa"]
                current_revenue = stats["summa"]["oylik_umumiy_summa"]
                if prev_revenue > 0:
                    revenue_pct = ((current_revenue - prev_revenue) / prev_revenue) * 100
                else:
                    revenue_pct = 100 if current_revenue > 0 else 0
                revenue_growth[month] = round(revenue_pct, 1)
            else:
                revenue_growth[month] = 0
            
            # GPS ulushi
            total_gps = stats["gps"]["jami_sotilgan"] + stats["gps"]["hozir_skladda_bor"]
            if total_gps > 0:
                gps_penetration[month] = round((stats["gps"]["jami_sotilgan"] / total_gps) * 100, 1)
            else:
                gps_penetration[month] = 0
            
            # To'lovlar nisbati
            total_subscribers = stats["tolov"]["oylik_tolaganlar"] + stats["tolov"]["oylik_tolamaganlar"]
            if total_subscribers > 0:
                payment_ratios[month] = round((stats["tolov"]["oylik_tolaganlar"] / total_subscribers) * 100, 1)
            else:
                payment_ratios[month] = 0
        
        # Eng ko'p sotilgan oylarni aniqlash
        monthly_sales = [(month, stats["gps"]["oylik_sotilgan"]) for month, stats in monthly_stats.items()]
        monthly_sales.sort(key=lambda x: x[1], reverse=True)
        top_months = [self.oylar[m-1] for m, _ in monthly_sales[:3]]
        
        # Yillik statistika
        annual_totals = {
            "total_subscribers": monthly_stats[current_month]["abonent"]["jami_aktiv"] if current_month in monthly_stats else 0,
            "total_revenue": sum(stats["summa"]["oylik_umumiy_summa"] for stats in monthly_stats.values()),
            "total_gps_sold": monthly_stats[current_month]["gps"]["jami_sotilgan"] if current_month in monthly_stats else 0,
            "avg_monthly_growth": sum(growth_rates.values()) / len(growth_rates) if growth_rates else 0,
            "highest_month": self.oylar[monthly_sales[0][0]-1] if monthly_sales else "",
            "payment_success_rate": sum(payment_ratios.values()) / len(payment_ratios) if payment_ratios else 0,
        }
        
        return {
            "growth_rates": growth_rates,
            "revenue_growth": revenue_growth,
            "gps_penetration": gps_penetration,
            "payment_ratios": payment_ratios,
            "top_months": top_months,
            "annual_totals": annual_totals,
            "colors": {
                "primary": "rgba(59, 130, 246, 0.7)",
                "secondary": "rgba(249, 115, 22, 0.7)",
                "accent": "rgba(168, 85, 247, 0.7)",
                "success": "rgba(34, 197, 94, 0.7)",
                "error": "rgba(239, 68, 68, 0.7)",
                "info": "rgba(14, 165, 233, 0.7)"
            }
        }

    def get_optimized_stats(self, year, current_month):
        """
        Barcha statistika ma'lumotlarini minimal SQL so'rovlar bilan hisoblash.

        Bu funksiya quyidagilarni amalga oshiradi:
        1. Barcha modellar uchun minimal so'rovlar (har bir model uchun maksimal 1-2 so'rov)
        2. select_related va prefetch_related dan maksimal foydalanish
        3. Barcha oylar uchun ma'lumotlarni bir so'rov bilan olish
        4. Samarali annotate va aggregate'dan foydalanish
        """
        # Natija obyektini tayyorlash
        result = {}

        # O'tgan va joriy yil/oy ma'lumotlarini aniqlash
        prev_year = year - 1

        # ========== 1. UMUMMIY SO'ROVLAR BO'YICHA MA'LUMOTLAR ==========

        # 1.1. Sotish modeli uchun ma'lumotlar (1-SQL so'rov)
        # O'tgan yil va joriy yilning barcha oylaridagi sotish ma'lumotlari
        all_sales = (
            Sotish.objects.filter(
                models.Q(
                    sana__year=prev_year, sana__month=12
                )  # O'tgan yilning oxirgi oyi
                | models.Q(
                    sana__year=year, sana__month__lte=current_month
                )  # Joriy yilning barcha oylari
            )
            .select_related("dasturiy_taminot")  # Foreign key relationship
            .prefetch_related(
                "sklad_sotish_set", "sklad_sotish_set__sklad"  # Reverse relations
            )
            .order_by("sana__year", "sana__month")
        )

        # 1.2. Bugalteriya modeli uchun ma'lumotlar (1-SQL so'rov)
        # Joriy yilning barcha oylari uchun to'lov ma'lumotlari
        all_payments = (
            Bugalteriya.objects.filter(yil=year, oy__in=self.oylar[:current_month])
            .select_related("sotish", "gps")
            .order_by("oy")
        )

        # 1.3. Sklad modeli uchun ma'lumotlar (2-SQL so'rov)
        # Barcha sotilgan/sotilmagan GPS'lar
        unsold_gps = Sklad.objects.filter(sotildi_sotilmadi=False).count()

        # O'tgan yilning oxirgi oyida sotilmagan GPS'lar
        prev_year_unsold = Sklad.objects.filter(
            olingan_sana__year=prev_year,
            olingan_sana__month=12,
            sotildi_sotilmadi=False,
        ).count()

        # 1.4. Sklad_sotish orqali bog'langan GPS'lar (1-SQL so'rov)
        # Har bir o'yda sotilgan GPS'lar va umumiy sotilgan GPS'lar
        sklad_sotish_items = (
            Sklad_sotish.objects.filter(
                sotish__sana__year__lte=year, sotish__sana__month__lte=current_month
            )
            .select_related("sklad", "sotish")
            .order_by("sotish__sana")
        )

        # ========== 2. MA'LUMOTLARNI XOTIRADA QAYTA ISHLASH ==========

        # 2.1. Sotish ma'lumotlarini oylar bo'yicha guruhlash
        sales_by_month = {}
        cumulative_sales_count = 0
        prev_year_sales_count = 0

        for sale in all_sales:
            sale_year = sale.sana.year
            sale_month = sale.sana.month

            # O'tgan yil uchun alohida hisob
            if sale_year == prev_year:
                prev_year_sales_count += 1
                continue

            # Oylik va cumulative hisoblar
            if sale_month not in sales_by_month:
                sales_by_month[sale_month] = {
                    "count": 0,
                    "total_summa": 0,
                    "sim_count": 0,
                    "masterlik": 0,  # Masterlik xarajatlari
                }

            # Ushbu oy uchun hisoblarni yuritish
            sales_by_month[sale_month]["count"] += 1
            sales_by_month[sale_month]["total_summa"] += sale.summasi
            sales_by_month[sale_month]["masterlik"] += sale.master_summasi

            # SIM karta sonini hisoblash (vergullar bilan ajratilgan)
            sim_count = 1
            if sale.sim_karta and "," in sale.sim_karta:
                sim_count = sale.sim_karta.count(",") + 1
            sales_by_month[sale_month]["sim_count"] += sim_count

            # Jamlanma hisob
            cumulative_sales_count += 1

        # 2.2. To'lovlar ma'lumotlarini guruhlash
        payments_by_month = {}
        for month_num, month_name in enumerate(self.oylar[:current_month], 1):
            payments_by_month[month_num] = {
                "paid": 0,
                "unpaid": 0,
                "abonent_payments": 0,
            }

        # To'lovlar ma'lumotlarini olish
        for payment in all_payments:
            month_num = self.oylar.index(payment.oy) + 1
            if payment.abonent_tolov:
                payments_by_month[month_num]["paid"] += 1
                payments_by_month[month_num][
                    "abonent_payments"
                ] += payment.sotish.abonent_tulov
            else:
                payments_by_month[month_num]["unpaid"] += 1

        # 2.3. Sotilgan GPS'larni oylar bo'yicha guruhlash
        sold_gps_by_month = {}
        sold_gps_ids_by_month = {month: set() for month in range(1, current_month + 1)}
        all_sold_gps_ids = set()

        for item in sklad_sotish_items:
            sale_year = item.sotish.sana.year
            sale_month = item.sotish.sana.month

            # Faqat joriy yilga tegishli ma'lumotlarni olish
            if sale_year == year and sale_month <= current_month:
                sold_gps_ids_by_month[sale_month].add(item.sklad.id)

            # Barcha sotilgan GPS'lar
            all_sold_gps_ids.add(item.sklad.id)

        # Har bir oy uchun sotilgan GPS'lar soni
        for month in range(1, current_month + 1):
            monthly_sold = len(sold_gps_ids_by_month[month])

            # Oyning oxirigacha sotilgan barcha GPS'lar
            cumulative_sold = set()
            for m in range(1, month + 1):
                cumulative_sold.update(sold_gps_ids_by_month[m])

            sold_gps_by_month[month] = {
                "monthly": monthly_sold,
                "cumulative": len(cumulative_sold),
            }

        # ========== 3. NATIJANI SHAKLLANTIRISH ==========

        # Har bir oy uchun statistika ma'lumotlarini yig'ish
        for month in range(1, current_month + 1):
            prev_month = 12 if month == 1 else month - 1

            # Abonent statistikasi
            abonent_stats = {
                "oldingidagi_aktiv": (
                    prev_year_sales_count
                    if month == 1
                    else sales_by_month.get(prev_month, {}).get("count", 0)
                ),
                "oylik_qoshilgan": sales_by_month.get(month, {}).get("count", 0),
                "jami_aktiv": sum(
                    sales_by_month.get(m, {}).get("count", 0)
                    for m in range(1, month + 1)
                ),
            }

            # To'lovlar statistikasi
            tolov_stats = {
                "oylik_tolamaganlar": payments_by_month.get(month, {}).get("unpaid", 0),
                "oylik_tolaganlar": payments_by_month.get(month, {}).get("paid", 0),
            }

            # GPS qurilmalari statistikasi
            gps_stats = {
                "otgan_oy_sklad": prev_year_unsold if month == 1 else 0,
                "hozir_skladda_bor": unsold_gps,
                "oylik_sotilgan": sold_gps_by_month.get(month, {}).get("monthly", 0),
                "jami_sotilgan": sold_gps_by_month.get(month, {}).get("cumulative", 0),
            }

            # SIM karta statistikasi
            sim_stats = {
                "oylik": sales_by_month.get(month, {}).get("sim_count", 0),
            }

            # Summalar statistikasi
            summa_stats = {
                "oylik_umumiy_summa": sales_by_month.get(month, {}).get(
                    "total_summa", 0
                ),
                "oylik_abonent": payments_by_month.get(month, {}).get(
                    "abonent_payments", 0
                ),
                "oylik_masterlik": sales_by_month.get(month, {}).get("masterlik", 0),
                "sof_foyda": (
                    sales_by_month.get(month, {}).get("total_summa", 0) - 
                    sales_by_month.get(month, {}).get("masterlik", 0)
                ),
            }

            # Barcha ma'lumotlarni birlashtirish
            result[month] = {
                "abonent": abonent_stats,
                "tolov": tolov_stats,
                "gps": gps_stats,
                "sim": sim_stats,
                "summa": summa_stats,
                "oy": self.oylar[month - 1],
            }

        return result


class GPSAddExcelView(LoginRequiredMixin, View):
    template_name = "sklad.html"

    def get(self, request):
        # Excel shablonini yaratish
        if request.GET.get("download_template"):
            # Excel fayl yaratish
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GPS Ma'lumotlari"

            # Ustun nomlari
            headers = [
                "gps_id",
                "olingan_odam",
                "tel_raqam",
                "summa_prixod",
                "olingan_sana",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            # Namuna ma'lumotlar
            sample_data = [
                ["GPS001", "Ali Valiyev", "+998901234567", 100000, "2025-01-27"],
                ["GPS002", "Vali Aliyev", "+998907654321", 150000, "2025-01-27"],
            ]

            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value

            # Ustun kengliklari
            ws.column_dimensions["A"].width = 15  # gps_id
            ws.column_dimensions["B"].width = 25  # olingan_odam
            ws.column_dimensions["C"].width = 15  # tel_raqam
            ws.column_dimensions["D"].width = 15  # summa_prixod
            ws.column_dimensions["E"].width = 15  # olingan_sana

            # Excel faylni saqlash va yuborish
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = "attachment; filename=gps_template.xlsx"
            wb.save(response)
            return response

        return render(request, self.template_name, {"skladlist": Sklad.objects.all()})

    def post(self, request):
        try:
            excel_file = request.FILES.get("excel_file")
            if not excel_file:
                messages.error(request, "Excel fayl tanlanmagan!")
                return redirect("sklad-list")

            # Excel faylni tekshirish
            if not excel_file.name.endswith((".xlsx", ".xls")):
                messages.error(
                    request,
                    "Noto'g'ri fayl formati. Faqat .xlsx yoki .xls fayllar qabul qilinadi!",
                )
                return redirect("sklad-list")

            # Excel faylni o'qish
            df = pd.read_excel(excel_file)
            required_columns = [
                "gps_id",
                "olingan_odam",
                "tel_raqam",
                "summa_prixod",
                "olingan_sana",
            ]

            # Ustunlar mavjudligini tekshirish
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messages.error(
                    request,
                    f"Excel faylda quyidagi ustunlar mavjud emas: {', '.join(missing_columns)}",
                )
                return redirect("sklad-list")

            success_count = 0
            error_count = 0
            errors = []

            # Excel fayldagi ma'lumotlarni bazaga saqlash
            for index, row in df.iterrows():
                try:
                    # Bo'sh qatorlarni o'tkazib yuborish
                    if pd.isna(row["gps_id"]) or str(row["gps_id"]).strip() == "":
                        continue

                    # GPS ID ni tekshirish
                    gps_id = str(row["gps_id"]).strip()
                    if Sklad.objects.filter(gps_id=gps_id).exists():
                        errors.append(f"GPS ID {gps_id} allaqachon mavjud!")
                        error_count += 1
                        continue

                    # Ma'lumotlarni tozalash
                    olingan_odam = (
                        str(row["olingan_odam"]).strip()
                        if not pd.isna(row["olingan_odam"])
                        else ""
                    )
                    tel_raqam = (
                        str(row["tel_raqam"]).strip()
                        if not pd.isna(row["tel_raqam"])
                        else ""
                    )

                    try:
                        summa_prixod = float(row["summa_prixod"])
                    except (ValueError, TypeError):
                        summa_prixod = 0

                    try:
                        olingan_sana = pd.to_datetime(row["olingan_sana"]).date()
                    except:
                        olingan_sana = timezone.now().date()

                    # Yangi GPS qo'shish
                    Sklad.objects.create(
                        gps_id=gps_id,
                        olingan_odam=olingan_odam,
                        tel_raqam=tel_raqam,
                        summa_prixod=summa_prixod,
                        olingan_sana=olingan_sana,
                        sotildi_sotilmadi=False,
                    )
                    success_count += 1

                except Exception as e:
                    errors.append(f"Qator {index + 2}: {str(e)}")
                    error_count += 1

            # Natijalarni xabar qilish
            if success_count > 0:
                messages.success(
                    request, f"{success_count} ta GPS muvaffaqiyatli qo'shildi!"
                )
            if error_count > 0:
                messages.error(request, f"{error_count} ta xatolik yuz berdi!")
                for error in errors:
                    messages.error(request, error)

        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")

        return redirect("sklad-list")
